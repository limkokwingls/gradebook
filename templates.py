from openpyxl.worksheet.worksheet import Worksheet
from model import Module, Student
from rich import print
from sample_data import student_data
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.cell.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.styles import DEFAULT_FONT

from pick_utils import EXIT_LABEL
border = Side(border_style="thin", color="000000")


def get_classes(students: list[Student]):
    list = []
    for it in students:
        if it.class_name not in list:
            list.append(it.class_name)
    return list


def add_standard_data(sheet: Worksheet, module: Module):
    sheet["D2"] = "STUDENT MARK-SHEET"
    title: Cell = sheet["D2"]
    title.font = Font(b=True, size=12)
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.border = Border(
        top=border, left=border, right=border, bottom=border)
    sheet.merge_cells("D2:J2")

    sheet['A9'] = "Lecture's Name"
    sheet['A10'] = "Thabo Lebese"
    sheet.merge_cells("A9:C9")
    sheet.merge_cells("A10:C10")


def create_workbook(module: Module, classes: list[str]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for it in classes:
        sheet = workbook.create_sheet(it)
        add_standard_data(sheet, module)
    return workbook


def apply_default_styles(workbook: Workbook):
    font = Font(name="Arial")
    for sheet in workbook:
        sheet.row_dimensions[2].height = 40
        for row in sheet:
            for cell in row:
                cell.font = font


def main(module: Module):
    list = get_classes(student_data)
    workbook = create_workbook(module, list)
    apply_default_styles(workbook)
    workbook.save("__module_name__.xlsx")


if __name__ == '__main__':
    module = Module(
        id='101',
        code='PROG101',
        name='Principles of Programming Logic and Design',
    )
    main(module)
